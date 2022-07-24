"""
Parsing text files with TextFSM and export to SQLite/Excel with configuration file

Version: 2022.07.25
"""

import argparse
import ast
import os
import re
import sqlite3
import time
from logging import info, warning, error, basicConfig, FileHandler, StreamHandler
from io import StringIO
from openpyxl.styles import Color, PatternFill
from openpyxl import Workbook, load_workbook
import textfsm
import yaml

# LOG OPTIONS
basicConfig(
    handlers=[
        FileHandler('parse2excel_LOG.txt'),
        StreamHandler()
    ],
    format='%(asctime)s.%(msecs)03d %(levelname)s : %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level='INFO'
)


class Parsev:
    """ Class for general defs """

    # security problem for "exec" ! restrict import library !
    @staticmethod
    def text_to_exec(text):
        """ execute given text code """
        text_parse = ast.parse(text)
        return exec(compile(text_parse, "", mode="exec"))

    @staticmethod
    def all_element_to_str(input_list) -> list:
        """ convert all element to string for list-in-list """
        result = []
        for i in input_list:
            result.append([str(x) for x in i])
        return result

    @staticmethod
    def list_to_sql(input_list, headerlist, dbname, tablename='',
                    deletebeforedb=False, deletebeforetable=False):
        """ usage: list_to_sql([['mylist1','mylist2','mylist3']],['h1','h2','h3'],'testdb','testtable',True,True) """

        if deletebeforedb is True:
            try:
                os.remove(dbname + '.sqlite3')
            except Exception:
                pass

        con = sqlite3.connect(dbname+".sqlite3")
        cur = con.cursor()

        if deletebeforetable is True:
            cur.execute("drop table if exists "+tablename)

        if tablename == '':
            tablename = (
                str(headerlist)[1: -1]).replace(
                ",", "_").replace(
                "'", "").replace(
                " ", "")

        q = "create table if not exists " + \
            tablename+"("+str(headerlist)[1:-1]+")"
        cur.execute(q)

        list_element_len = len(input_list[0])
        question_mark = ('?,' * list_element_len)[:-1]
        with con:
            q = "insert into " + tablename + " values (" + question_mark + ")"
            cur.executemany(q, input_list)

    @staticmethod
    def dbtable_to_xlsx(db, table, file, sheetname, deletebefore=False):
        """ convert sqlite to excel file """
        if deletebefore:
            try:
                os.remove(file + '.xlsx')
            except Exception:
                pass

        con = sqlite3.connect(db)
        cur = con.cursor()
        # data values
        tablefetch = cur.execute('select * from ' + table).fetchall()
        # header
        headers = cur.execute('PRAGMA table_info(' + table + ')').fetchall()
        headerslist = []
        headerslist = [headers[i][1] for i, h in enumerate(headers)]

        if os.path.isfile(file + '.xlsx'):
            wb = load_workbook(filename=file + '.xlsx', data_only=True)
        else:
            wb = Workbook()
            ws = wb.active
            wb.remove(ws)
        if sheetname == '':
            ws = wb.create_sheet(title=f'{db}-{table}')
            ws.freeze_panes = 'A2'
        else:
            ws = wb.create_sheet(title=sheetname)
            ws.freeze_panes = 'A2'
        # add header and table to ws
        ws.append(headerslist)
        for row in tablefetch:
            ws.append(row)
        # add filter
        ws.auto_filter.ref = ws.dimensions
        # style header fill
        for i in range(1, len(headers) + 1):
            ws.cell(row=1, column=i).fill = PatternFill(
                patternType='solid', fgColor=Color('FFFF00'))
        for i in range(ord('a'), ord('z') + 1):
            ws.column_dimensions[chr(i)].width = 25.0

        wb.save(filename=file + '.xlsx')

    @staticmethod
    def yaml_file_to_list(yaml_file) -> list:
        """ convert yaml file to list"""
        with open(yaml_file, encoding='utf-8') as file:
            yaml_text = file.read()
        try:
            all_parts = list(yaml.full_load_all(yaml_text))[0]
            return all_parts
        except Exception as e:
            error(f'CHECK YAML FORMAT : {e}')
            return []


class Textfsmv:
    """ Textfsm based class """

    @staticmethod
    def _textfsm_result_with_host(
            output, textfsm_template, host='', isfile=False):
        """textfsm_result with hostname"""
        if isfile:
            template = open(textfsm_template)
            fsm_result = textfsm.TextFSM(template).ParseText(output)
        else:
            try:
                template = StringIO(textfsm_template)
                fsm_result_header = textfsm.TextFSM(template).header
                fsm_result = textfsm.TextFSM(template).ParseText(output)
                fsm_result_host = [[host]+i for i in fsm_result]
                result = [['Filename'] + fsm_result_header] + fsm_result_host
                return result
            except Exception as e:
                error(f'check textfsm template : {e}')
                raise SystemExit

    @staticmethod
    def textfsmv_run_yaml(yaml_file, timestamp, excel_export=True):
        """run with textfsm yaml file"""
        all_parts = [i for i in Parsev.yaml_file_to_list(
            yaml_file) if i['type'] == 'textfsm']

        for part in all_parts:
            excel_export = True
            # textfsm template from yaml
            textfsm_input = part['template']
            # db name from yaml
            sql_dbname = part['db_name'] + timestamp
            # table name from yaml AND replace some chars
            table_name = part['table_name'].replace('.', '_').replace('-', '_')
            if 'excel_export' in part:
                if part['excel_export'].lower() == 'none':
                    excel_export = False

            # for single hosts, yaml:files
            if 'files' in part:
                for host_file in part['files']:
                    with open(host_file, encoding='utf-8') as file:
                        host_file_text = file.read()
                    # textfsm_result to sqlite
                    single_textfsm_result = Textfsmv._textfsm_result_with_host(
                        host_file_text, textfsm_input, host_file.split('/')[-1])
                    try:
                        if len(single_textfsm_result) < 2:
                            warning(
                                f'check textfsm template NO TEXTFSM RESULT! @ TABLE: {table_name} HOST: {host_file}')
                            continue
                        single_textfsm_result_header = single_textfsm_result[0]
                        # single_textfsm_result_data = single_textfsm_result[1:]
                        # convert list element to string
                        single_textfsm_result_data = Parsev.all_element_to_str(
                            single_textfsm_result[1:])
                        Parsev.list_to_sql(
                            single_textfsm_result_data,
                            single_textfsm_result_header, sql_dbname,
                            table_name)
                    except Exception as e:
                        error(f'list to sql problem : {e}')

            # for folders yaml:folders
            if 'folders' in part:
                for folder in part['folders']:
                    all_host_file = os.listdir(folder + "/")
                    for host_file in all_host_file:
                        with open(folder + '/' + host_file, encoding='utf-8') as file:
                            host_file_text = file.read()
                            # textfsm_result to sqlite
                            single_textfsm_result = Textfsmv._textfsm_result_with_host(
                                host_file_text, textfsm_input, host_file)
                            try:
                                if len(single_textfsm_result) < 2:
                                    warning(
                                        f'check textfsm template NO TEXTFSM RESULT! @ TABLE: {table_name} HOST: {host_file}')
                                    continue
                                single_textfsm_result_header = single_textfsm_result[0]
                                # single_textfsm_result_data = single_textfsm_result[1:]
                                # convert list element to string
                                single_textfsm_result_data = Parsev.all_element_to_str(
                                    single_textfsm_result[1:])
                                Parsev.list_to_sql(
                                    single_textfsm_result_data,
                                    single_textfsm_result_header, sql_dbname,
                                    table_name)
                            except Exception as e:
                                error(f'list to sql problem : {e}')

            info(f'[{table_name}] / [{sql_dbname}] SQL COMPLETED!')
            # excel export
            if excel_export:
                try:
                    Parsev.dbtable_to_xlsx(
                        sql_dbname + '.sqlite3', table_name, sql_dbname,
                        table_name)
                    info(f'[{table_name}] / [{sql_dbname}] EXCEL SHEET COMPLETED!')
                except Exception as e:
                    error(f'sql to excel problem : {e}')

            if 'files' not in part and 'folders' not in part:
                error('no files or folder in yaml')


class Sqljoinv:
    """ Sqljoin class """
    @staticmethod
    def _get_functions_from_config(yaml_file):
        functions_text = ''
        all_parts = [i for i in Parsev.yaml_file_to_list(
            yaml_file) if i['type'] == 'sqlfunction']

        for part in all_parts:
            if 'functions' in part:
                for func in part['functions']:
                    functions_text += func + '\n'

        return functions_text

    @staticmethod
    def sqljoinv_run_yaml(yaml_file, timestamp, excel_export=True):
        """run with sqljoin yaml file"""
        all_parts = [i for i in Parsev.yaml_file_to_list(
            yaml_file) if i['type'] == 'sqljoin']

        for part in all_parts:
            excel_export = True
            try:
                sql_dbname = part['db_name'] + timestamp
                if 'excel_export' in part:
                    if part['excel_export'].lower() == 'none':
                        excel_export = False
                if 'sqlcommand' in part:
                    sqlcommand = part['sqlcommand']
                    new_table = part['new_table']

                    # if sqlfunction type in config
                    sqlfunction_text = Sqljoinv._get_functions_from_config(
                        yaml_file).strip()
                    if sqlfunction_text != '':
                        # split with "^def" regex and strip
                        func_files_list = [
                            'def ' + i
                            for i in re.split(
                                '^def ', sqlfunction_text,
                                flags=re.MULTILINE) if i]
                        for func in func_files_list:
                            # below lines same as non-file function
                            func_name = func.split('def ')[1].split('(')[0]

                            ###
                            # add try/except to custom function
                            func_try_prefix = "  try:"
                            func_try_suffix = "  except Exception as e:\n    print('check custom function @ " + \
                                func_name + " : ' + str(e))"
                            # new func without FIRST 'def x(..):' line
                            new_func_wo_def = '\n'.join(
                                [f'  {i}' for i in func.splitlines()[1:]])
                            first_def_line = func.splitlines()[0]
                            if 'def ' not in first_def_line:
                                error(
                                    f'check custom function @ {func_name}: custom function not started with <def >')
                                raise Exception(
                                    f'check custom function @ {func_name}: custom function not started with <def >')
                            new_func = f'{first_def_line}\n{func_try_prefix}\n{new_func_wo_def}\n{func_try_suffix}\n'
                            ###

                            # adding global to def
                            global_text = f"global {func_name}\n"
                            # exec with new function text
                            def_text_exec = global_text + new_func
                            Parsev.text_to_exec(def_text_exec)

                    # if function exist
                    if 'functions' in part:
                        for func in part['functions']:
                            func_name = func.split('def ')[1].split('(')[0]

                            ###
                            # add try/except to custom function
                            func_try_prefix = "  try:"
                            func_try_suffix = "  except Exception as e:\n    print('check custom function @ " + \
                                func_name + " : ' + str(e))"
                            # new func without FIRST 'def x(..):' line
                            new_func_wo_def = '\n'.join(
                                [f'  {i}' for i in func.splitlines()[1:]])
                            first_def_line = func.splitlines()[0]
                            if 'def ' not in first_def_line:
                                error(
                                    f'check custom function @ {func_name}: custom function not started with <def >')
                                raise Exception(
                                    f'check custom function @ {func_name}: custom function not started with <def >')
                            new_func = f'{first_def_line}\n{func_try_prefix}\n{new_func_wo_def}\n{func_try_suffix}\n'
                            ###

                            # adding global to def
                            global_text = f"global {func_name}\n"
                            # exec with new function text
                            def_text_exec = global_text + new_func
                            Parsev.text_to_exec(def_text_exec)

                elif 'sqlcommand_run' in part:
                    sqlcommand = part['sqlcommand_run']
                    new_table = 'CUSTOM SQL COMMAND RUN'
                    excel_export = False

                # auto left join
                else:
                    first_table = part['first_table']
                    second_table = part['second_table']
                    match = part['match']
                    new_table = part['new_table']

                    if '=' in match:
                        sqlcommand = (
                            f'SELECT * FROM {first_table} '
                            f'LEFT OUTER JOIN {second_table} '
                            f'ON {match} '
                        )
                    else:
                        match_list = [i.strip() for i in match.split(',')]
                        match_list.append('Filename')
                        on_command = ''
                        for m in match_list:
                            on_command += f'{first_table}.{m} = {second_table}.{m} AND '
                        on_command = on_command[:-5]
                        sqlcommand = (
                            f'SELECT * FROM {first_table} '
                            f'LEFT OUTER JOIN {second_table} '
                            f'ON {on_command}'
                        )
                #

                with sqlite3.connect(sql_dbname+".sqlite3") as con:
                    cur = con.cursor()
                    # if sqlfunction type in config
                    sqlfunction_text = Sqljoinv._get_functions_from_config(
                        yaml_file).strip()
                    if sqlfunction_text != '':
                        # split with "^def" regex and strip
                        func_files_list = [
                            'def ' + i
                            for i in re.split(
                                '^def ', sqlfunction_text,
                                flags=re.MULTILINE) if i]
                        for func in func_files_list:
                            # below lines same as non-file function
                            func_name = func.split('def ')[1].split('(')[0]
                            con.create_function(
                                func_name,
                                globals()[func_name].__code__.co_argcount,
                                globals()[func_name])
                    # if function exist
                    if 'functions' in part:
                        for func in part['functions']:
                            func_name = func.split('def ')[1].split('(')[0]
                            con.create_function(
                                func_name,
                                globals()[func_name].__code__.co_argcount,
                                globals()[func_name])

                    # Run raw sqlite3 command if sqlcommand_run in config
                    if 'sqlcommand_run' in part:
                        cur.execute(f"{sqlcommand}")
                        info(
                            f'SQL COMMAND: {sqlcommand} RESULT: {cur.fetchall()}')
                    else:
                        cur.execute(f"drop table if exists {new_table}")
                        cur.execute(
                            f"CREATE TABLE {new_table} AS {sqlcommand}")

            except Exception as e:
                error(f'check sqljoin @ {part} : {e}')
                # continue other parts
                continue

            info(f'[{new_table}] / [{sql_dbname}] SQL COMPLETED!')
            # excel export
            if excel_export:
                try:
                    Parsev.dbtable_to_xlsx(
                        sql_dbname + '.sqlite3', new_table, sql_dbname,
                        new_table)
                    info(f'[{new_table}] / [{sql_dbname}] EXCEL SHEET COMPLETED!')
                except Exception as e:
                    error(f'sql to excel problem : {e}')


class Excel2Sql:
    """ Excel file to Sqlite """
    @staticmethod
    def _get_excel_row(
            excel_sheet, row_start=None, row_end=None, col_start=None,
            col_end=None):
        ''' Strip cell and remove none values, excel_sheet is worksheet in openpyxl workbook '''

        table_strip = []
        for row in excel_sheet.iter_rows(
                min_row=row_start, max_row=row_end, min_col=col_start,
                max_col=col_end, values_only=True):
            # 'None' to ''
            row_strip_wo_none = []
            for i in row:
                if str(i) == 'None':
                    row_strip_wo_none.append('')
                else:
                    row_strip_wo_none.append(str(i).strip())
            table_strip.append(row_strip_wo_none)
        return table_strip

    @staticmethod
    def excel_to_sql(excel_file, excel_sheets=None, db_name=None):
        """ Convert excel file to sqlite file: excel_sheets is list (['sheetname1', 'sheetname2']) OR None for all sheets """
        wb = load_workbook(filename=excel_file, data_only=True)
        if not db_name:
            db_name = excel_file.split('.xl')[0]

        for sheet in wb:
            sheet_name = sheet.title
            # check excel_sheets specify
            if excel_sheets and sheet_name in excel_sheets:
                sheet = wb[sheet_name]
            header = Excel2Sql._get_excel_row(
                sheet, row_start=1, col_start=1)[0]
            header_wo_none = [i for i in header if i != '']
            data = Excel2Sql._get_excel_row(sheet, row_start=2, col_start=1)
            Parsev.list_to_sql(data, header_wo_none, db_name, sheet_name)

    @staticmethod
    def excel_run_yaml(yaml_file, timestamp):
        """ run with sqljoin yaml file """
        all_parts = [i for i in Parsev.yaml_file_to_list(
            yaml_file) if i['type'] == 'excel']

        for part in all_parts:
            try:
                sql_dbname = part['db_name'] + timestamp
                excel_file = part['excel_file']
                excel_sheets = None
                # check excel_sheets
                if 'excel_sheets' in part:
                    excel_sheets = part['excel_sheets']
                Excel2Sql.excel_to_sql(
                    excel_file, excel_sheets=excel_sheets, db_name=sql_dbname)
                info(f'[{excel_file}] EXCEL TO SQL COMPLETED!')
            except Exception as e:
                error(f'check excel @ {part} : {e}')
                raise SystemExit


def main():
    ''' main function to run parse2excel '''
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'configfile',
        help='config yaml file path [e.g. srlinux_config_1.yaml] (OPTIONAL default: file=config.yaml, folder=P2E_CONFIGS)',
        nargs='?')
    args = parser.parse_args()
    if args.configfile:
        config_file_path = args.configfile
    else:
        config_file_path = 'config.yaml'

    if os.path.exists(config_file_path):
        try:
            timestamp = '_' + time.strftime("%Y%m%d-%H%M%S")
            info(f'START <{config_file_path}> CONFIG FILE!')
            Excel2Sql.excel_run_yaml(config_file_path, timestamp)
            Textfsmv.textfsmv_run_yaml(config_file_path, timestamp)
            Sqljoinv.sqljoinv_run_yaml(config_file_path, timestamp)
            info('ALL DONE!')
            input('!!! ALL DONE! Press any key to exit...')
        except Exception as e:
            error(
                f'NOT DONE! CHECK ERRORS AT <textfsm_CONFIG.yaml> CONFIG FILE! : {e}')
            input('!!! NOT DONE! CHECK ERRORS! Press any key to exit...')
    elif os.path.exists('P2E_CONFIGS'):
        for config_file in os.listdir('P2E_CONFIGS'):
            config_file_path = os.path.join('P2E_CONFIGS', config_file)
            try:
                timestamp = '_' + time.strftime("%Y%m%d-%H%M%S")
                info(f'START <{config_file}> CONFIG FILE!')
                Excel2Sql.excel_run_yaml(config_file_path, timestamp)
                Textfsmv.textfsmv_run_yaml(config_file_path, timestamp)
                Sqljoinv.sqljoinv_run_yaml(config_file_path, timestamp)
                info('ALL DONE!')
                input('!!! ALL DONE! Press any key to exit...')
            except Exception as e:
                error(
                    f'NOT DONE! CHECK ERRORS AT <{config_file}> CONFIG FILE! : {e}')
                input('!!! NOT DONE! CHECK ERRORS! Press any key to exit...')
    else:
        input(
            '!!! NOT DONE! <textfsm_CONFIG.yaml> file not found! Press any key to exit...')


if __name__ == "__main__":
    main()
